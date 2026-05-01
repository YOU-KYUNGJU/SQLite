from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from copy import deepcopy
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


DEFAULT_CONFIG = {
    "sheet_name": "접수내역",
    "table_name": "receipt_status",
    "db_mode": "single_db",
    "table_mode": "fixed",
    "header_row": 3,
    "data_start_row": 4,
    "date_column": "접수일자",
    "pattern": "*.xlsx",
    "append_year_to_source_dir": True,
    "single_db_name_template": "receipt_status_{year}.db",
    "part_code": "default",
    "part_name": "기본",
    "part_description": "기본 파트 데이터",
}

HEADER_ALIASES = {
    "접수번호": "receipt_no",
    "접수일자": "received_date",
    "발급예정일": "due_date",
    "발급일": "issued_date",
    "발급구분": "issue_type",
    "접수부서": "department",
    "결과입력일": "result_input_date",
    "시험항목": "test_item",
    "규격": "test_method",
    "진행상황": "status",
    "신청업체": "client_name",
    "납품업체": "supplier_name",
    "Buyer": "buyer_name",
    "시료수": "sample_count",
    "항목수수료": "item_fee",
    "전체수수료": "total_fee",
    "Remark of Receptioin": "remark_reception",
    "Remark of Test": "remark_test",
    "Remark of Entry": "remark_entry",
}


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="엑셀 파일을 SQLite로 적재합니다."
    )
    parser.add_argument("--config", type=Path, help="설정 JSON 파일 경로")
    parser.add_argument("--source-dir", type=Path, help="원본 엑셀 기준 폴더")
    parser.add_argument("--target-dir", type=Path, help="DB 저장 폴더")
    parser.add_argument("--year", default=str(datetime.now().year), help="연도 폴더명")
    parser.add_argument("--pattern", help="원본 엑셀 검색 패턴")
    parser.add_argument("--sheet-name", help="적재할 시트명")
    parser.add_argument("--table-name", help="적재할 테이블명")
    parser.add_argument("--db-mode", choices=("per_file", "single_db"))
    parser.add_argument("--table-mode", choices=("fixed", "per_file"))
    parser.add_argument("--single-db-name", help="단일 DB 모드 파일명")
    parser.add_argument("--header-row", type=int)
    parser.add_argument("--data-start-row", type=int)
    parser.add_argument("--date-column", help="연/월 추출 기준 컬럼명")
    parser.add_argument("--part-code", help="파트 코드")
    parser.add_argument("--part-name", help="파트명")
    parser.add_argument("--part-description", help="파트 설명")
    parser.add_argument("--replace-db", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args(argv)


def sanitize_identifier(name: str) -> str:
    text = re.sub(r"[^0-9A-Za-z_]+", "_", name.strip())
    text = re.sub(r"_+", "_", text).strip("_").lower()
    if not text:
        raise ValueError("Identifier is empty after sanitization.")
    if text[0].isdigit():
        text = f"t_{text}"
    return text


def normalize_header_name(name: object, index: int) -> str:
    text = "" if name is None else str(name).strip()
    if not text:
        return f"column_{index}"
    if text in HEADER_ALIASES:
        return HEADER_ALIASES[text]
    sanitized = re.sub(r"[^0-9A-Za-z_]+", "_", text).strip("_").lower()
    if not sanitized:
        sanitized = f"column_{index}"
    if sanitized[0].isdigit():
        sanitized = f"col_{sanitized}"
    return sanitized


def excel_value_to_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value).strip()


def load_config(path: Path | None) -> dict[str, Any]:
    config = deepcopy(DEFAULT_CONFIG)
    if path is None:
        return config
    loaded = json.loads(path.read_text(encoding="utf-8"))
    config.update(loaded)
    return config


def apply_cli_overrides(config: dict[str, Any], args: argparse.Namespace) -> dict[str, Any]:
    merged = deepcopy(config)
    fields = {
        "source_dir": args.source_dir,
        "target_dir": args.target_dir,
        "year": args.year,
        "pattern": args.pattern,
        "sheet_name": args.sheet_name,
        "table_name": args.table_name,
        "db_mode": args.db_mode,
        "table_mode": args.table_mode,
        "single_db_name": args.single_db_name,
        "header_row": args.header_row,
        "data_start_row": args.data_start_row,
        "date_column": args.date_column,
        "part_code": args.part_code,
        "part_name": args.part_name,
        "part_description": args.part_description,
    }
    for key, value in fields.items():
        if value is not None:
            merged[key] = value
    merged["replace_db"] = bool(args.replace_db)
    merged["dry_run"] = bool(args.dry_run)
    return merged


def resolve_source_dir(config: dict[str, Any]) -> Path:
    year = str(config["year"])
    append_year = bool(config.get("append_year_to_source_dir", True))

    if config.get("source_dir"):
        base = Path(config["source_dir"])
        return base / year if append_year else base

    for source_base in config.get("source_base_dirs", []):
        candidate = Path(source_base)
        candidate = candidate / year if append_year else candidate
        print(f"Checking source directory: {candidate}")
        if candidate.exists():
            print(f"Using source directory: {candidate}")
            return candidate

    raise FileNotFoundError("No usable source directory was found from config.")


def resolve_target_dir(config: dict[str, Any]) -> Path:
    if not config.get("target_dir"):
        raise ValueError("target_dir is required.")
    return Path(config["target_dir"])


def resolve_single_db_name(config: dict[str, Any]) -> str:
    if config.get("single_db_name"):
        return str(config["single_db_name"])
    template = str(config["single_db_name_template"])
    return template.format(year=config["year"], part_code=config["part_code"])


def read_header(worksheet, header_row: int) -> tuple[list[str], list[str]]:
    raw_header = next(
        worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            values_only=True,
        )
    )
    last_nonempty_index = -1
    for index, value in enumerate(raw_header):
        if value not in (None, ""):
            last_nonempty_index = index
    if last_nonempty_index < 0:
        raise ValueError("Header row is empty.")

    source_headers = [
        "" if value is None else str(value).strip()
        for value in raw_header[: last_nonempty_index + 1]
    ]
    normalized_headers = [
        normalize_header_name(value, index)
        for index, value in enumerate(source_headers, start=1)
    ]
    return source_headers, normalized_headers


def parse_year_month(date_text: str | None) -> tuple[int | None, int | None]:
    if not date_text:
        return None, None
    match = re.match(r"^\s*(\d{4})-(\d{2})-(\d{2})", date_text)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None, None


def iter_data_rows(
    worksheet,
    data_start_row: int,
    column_count: int,
) -> Iterable[tuple[int, list[str | None]]]:
    for source_row, row in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True),
        start=data_start_row,
    ):
        values = [excel_value_to_text(value) for value in row[:column_count]]
        if any(value not in (None, "") for value in values):
            yield source_row, values


def ensure_parent_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def resolve_db_path(
    *,
    target_dir: Path,
    excel_path: Path,
    db_mode: str,
    single_db_name: str,
) -> Path:
    if db_mode == "per_file":
        return target_dir / f"{excel_path.stem}.db"
    return target_dir / single_db_name


def resolve_table_name(
    *,
    excel_path: Path,
    table_mode: str,
    table_name: str,
) -> str:
    if table_mode == "per_file":
        return sanitize_identifier(excel_path.stem)
    return sanitize_identifier(table_name)


def prepare_database_file(db_path: Path, replace_db: bool) -> None:
    ensure_parent_dir(db_path.parent)
    if db_path.exists():
        if replace_db:
            db_path.unlink()
        else:
            raise FileExistsError(
                f"Database already exists: {db_path}. Use --replace-db to overwrite it."
            )


def create_import_metadata_table(connection: sqlite3.Connection) -> None:
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS import_metadata (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_code TEXT NOT NULL,
            part_name TEXT NOT NULL,
            source_file TEXT NOT NULL,
            source_sheet TEXT NOT NULL,
            table_name TEXT NOT NULL,
            imported_at_utc TEXT NOT NULL,
            row_count INTEGER NOT NULL
        )
        """
    )


def create_file_import_logs_table(connection: sqlite3.Connection) -> None:
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS file_import_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_code TEXT NOT NULL,
            part_name TEXT NOT NULL,
            file_name TEXT NOT NULL,
            file_path TEXT NOT NULL,
            file_hash TEXT,
            imported_at TEXT NOT NULL,
            total_rows INTEGER NOT NULL,
            inserted_rows INTEGER NOT NULL,
            duplicated_rows INTEGER NOT NULL DEFAULT 0,
            failed_rows INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL,
            error_message TEXT
        )
        """
    )
    connection.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_file_import_logs_file_name
        ON file_import_logs (file_name)
        """
    )


def validate_existing_table(
    connection: sqlite3.Connection,
    *,
    table_name: str,
    expected_columns: list[str],
) -> None:
    rows = connection.execute(f'PRAGMA table_info("{table_name}")').fetchall()
    if not rows:
        return
    existing_columns = [row[1] for row in rows]
    if existing_columns != expected_columns:
        raise ValueError(
            f"Existing table schema does not match for {table_name}.\n"
            f"Expected: {expected_columns}\n"
            f"Actual: {existing_columns}"
        )


def create_receipt_status_table(connection: sqlite3.Connection, *, table_name: str) -> None:
    expected_columns = [
        "id",
        "part_code",
        "part_name",
        "part_description",
        "receipt_no",
        "test_item",
        "department",
        "sample_name",
        "client_name",
        "received_date",
        "due_date",
        "status",
        "issue_type",
        "issued_date",
        "result_input_date",
        "test_method",
        "supplier_name",
        "buyer_name",
        "sample_count",
        "item_fee",
        "total_fee",
        "import_year",
        "import_month",
        "source_file",
        "source_sheet",
        "source_row",
        "created_at",
        "updated_at",
        "remark",
        "remark_reception",
        "remark_test",
        "remark_entry",
    ]
    validate_existing_table(connection, table_name=table_name, expected_columns=expected_columns)
    connection.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{table_name}" (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_code TEXT NOT NULL,
            part_name TEXT NOT NULL,
            part_description TEXT,
            receipt_no TEXT NOT NULL,
            test_item TEXT,
            department TEXT,
            sample_name TEXT,
            client_name TEXT,
            received_date TEXT,
            due_date TEXT,
            status TEXT,
            issue_type TEXT,
            issued_date TEXT,
            result_input_date TEXT,
            test_method TEXT,
            supplier_name TEXT,
            buyer_name TEXT,
            sample_count TEXT,
            item_fee TEXT,
            total_fee TEXT,
            import_year INTEGER,
            import_month INTEGER,
            source_file TEXT NOT NULL,
            source_sheet TEXT NOT NULL,
            source_row INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            remark TEXT,
            remark_reception TEXT,
            remark_test TEXT,
            remark_entry TEXT
        )
        '''
    )
    for candidate in (
        "part_code",
        "part_name",
        "receipt_no",
        "status",
        "received_date",
        "source_file",
    ):
        index_name = sanitize_identifier(f"idx_{table_name}_{candidate}")
        connection.execute(
            f'CREATE INDEX IF NOT EXISTS "{index_name}" ON "{table_name}" ("{candidate}")'
        )
    compound_index = sanitize_identifier(f"idx_{table_name}_part_receipt_no")
    connection.execute(
        f'CREATE INDEX IF NOT EXISTS "{compound_index}" ON "{table_name}" ("part_code", "receipt_no")'
    )


def build_record(
    *,
    normalized_headers: list[str],
    row_values: list[str | None],
    source_file: str,
    source_sheet: str,
    source_row: int,
    created_at: str,
    config: dict[str, Any],
) -> dict[str, object]:
    row_map = dict(zip(normalized_headers, row_values))
    received_date = row_map.get("received_date")
    import_year, import_month = parse_year_month(
        None if received_date is None else str(received_date)
    )

    client_name = row_map.get("client_name")
    supplier_name = row_map.get("supplier_name")
    sample_name = client_name if client_name else supplier_name
    remark_parts = [
        row_map.get("remark_reception"),
        row_map.get("remark_test"),
        row_map.get("remark_entry"),
    ]
    remark = " | ".join(part for part in remark_parts if part)

    return {
        "part_code": config["part_code"],
        "part_name": config["part_name"],
        "part_description": config.get("part_description"),
        "receipt_no": row_map.get("receipt_no"),
        "test_item": row_map.get("test_item"),
        "department": row_map.get("department"),
        "sample_name": sample_name,
        "client_name": client_name,
        "received_date": received_date,
        "due_date": row_map.get("due_date"),
        "status": row_map.get("status"),
        "issue_type": row_map.get("issue_type"),
        "issued_date": row_map.get("issued_date"),
        "result_input_date": row_map.get("result_input_date"),
        "test_method": row_map.get("test_method"),
        "supplier_name": supplier_name,
        "buyer_name": row_map.get("buyer_name"),
        "sample_count": row_map.get("sample_count"),
        "item_fee": row_map.get("item_fee"),
        "total_fee": row_map.get("total_fee"),
        "import_year": import_year,
        "import_month": import_month,
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_row": source_row,
        "created_at": created_at,
        "updated_at": created_at,
        "remark": remark or None,
        "remark_reception": row_map.get("remark_reception"),
        "remark_test": row_map.get("remark_test"),
        "remark_entry": row_map.get("remark_entry"),
    }


def compute_file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_handle:
        for chunk in iter(lambda: file_handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def import_workbook(
    *,
    excel_path: Path,
    db_path: Path,
    sheet_name: str,
    table_name: str,
    header_row: int,
    data_start_row: int,
    date_column: str,
    dry_run: bool,
    config: dict[str, Any],
) -> tuple[Path, str, int]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' not found in {excel_path.name}.")

    worksheet = workbook[sheet_name]
    _, normalized_headers = read_header(worksheet, header_row)
    date_alias = normalize_header_name(date_column, 0)
    if date_alias not in normalized_headers:
        raise ValueError(
            f"Date column '{date_column}' not found in {excel_path.name}."
        )

    row_count = sum(1 for _ in iter_data_rows(worksheet, data_start_row, len(normalized_headers)))
    if dry_run:
        workbook.close()
        return db_path, table_name, row_count

    workbook.close()
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    _, normalized_headers = read_header(worksheet, header_row)
    imported_at = datetime.now(UTC).isoformat()

    records: list[dict[str, object]] = []
    failed_rows = 0
    for source_row, row_values in iter_data_rows(worksheet, data_start_row, len(normalized_headers)):
        record = build_record(
            normalized_headers=normalized_headers,
            row_values=row_values,
            source_file=excel_path.name,
            source_sheet=sheet_name,
            source_row=source_row,
            created_at=imported_at,
            config=config,
        )
        if not record["receipt_no"]:
            failed_rows += 1
            continue
        records.append(record)

    file_hash = compute_file_hash(excel_path)

    with sqlite3.connect(db_path) as connection:
        connection.execute("PRAGMA journal_mode=WAL")
        create_import_metadata_table(connection)
        create_file_import_logs_table(connection)
        create_receipt_status_table(connection, table_name=table_name)
        connection.executemany(
            f'''
            INSERT INTO "{table_name}" (
                part_code,
                part_name,
                part_description,
                receipt_no,
                test_item,
                department,
                sample_name,
                client_name,
                received_date,
                due_date,
                status,
                issue_type,
                issued_date,
                result_input_date,
                test_method,
                supplier_name,
                buyer_name,
                sample_count,
                item_fee,
                total_fee,
                import_year,
                import_month,
                source_file,
                source_sheet,
                source_row,
                created_at,
                updated_at,
                remark,
                remark_reception,
                remark_test,
                remark_entry
            ) VALUES (
                :part_code,
                :part_name,
                :part_description,
                :receipt_no,
                :test_item,
                :department,
                :sample_name,
                :client_name,
                :received_date,
                :due_date,
                :status,
                :issue_type,
                :issued_date,
                :result_input_date,
                :test_method,
                :supplier_name,
                :buyer_name,
                :sample_count,
                :item_fee,
                :total_fee,
                :import_year,
                :import_month,
                :source_file,
                :source_sheet,
                :source_row,
                :created_at,
                :updated_at,
                :remark,
                :remark_reception,
                :remark_test,
                :remark_entry
            )
            ''',
            records,
        )
        connection.execute(
            """
            INSERT INTO import_metadata (
                part_code,
                part_name,
                source_file,
                source_sheet,
                table_name,
                imported_at_utc,
                row_count
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                config["part_code"],
                config["part_name"],
                excel_path.name,
                sheet_name,
                table_name,
                imported_at,
                len(records),
            ),
        )
        connection.execute(
            """
            INSERT INTO file_import_logs (
                part_code,
                part_name,
                file_name,
                file_path,
                file_hash,
                imported_at,
                total_rows,
                inserted_rows,
                duplicated_rows,
                failed_rows,
                status,
                error_message
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                config["part_code"],
                config["part_name"],
                excel_path.name,
                str(excel_path),
                file_hash,
                imported_at,
                row_count,
                len(records),
                0,
                failed_rows,
                "SUCCESS" if failed_rows == 0 else "PARTIAL",
                None if failed_rows == 0 else "Some rows were skipped because receipt_no was empty.",
            ),
        )
        connection.commit()

    workbook.close()
    return db_path, table_name, len(records)


def run_import(config: dict[str, Any]) -> int:
    source_dir = resolve_source_dir(config)
    target_dir = resolve_target_dir(config)
    excel_files = sorted(source_dir.glob(config["pattern"]))
    if not excel_files:
        raise FileNotFoundError(f"No Excel files matched '{config['pattern']}' in {source_dir}")

    if config["db_mode"] == "per_file" and config["table_mode"] == "per_file":
        raise ValueError("Per-file DB mode does not need per-file table mode.")

    print(f"Source directory: {source_dir}")
    print(f"Year: {config['year']}")
    print(f"Target directory: {target_dir}")
    print(f"Files found: {len(excel_files)}")
    print(f"Part: {config['part_name']} ({config['part_code']})")
    print(f"Database mode: {config['db_mode']}")
    print(f"Table mode: {config['table_mode']}")

    prepared_single_db = False
    single_db_name = resolve_single_db_name(config)

    for excel_file in excel_files:
        db_path = resolve_db_path(
            target_dir=target_dir,
            excel_path=excel_file,
            db_mode=config["db_mode"],
            single_db_name=single_db_name,
        )
        table_name = resolve_table_name(
            excel_path=excel_file,
            table_mode=config["table_mode"],
            table_name=config["table_name"],
        )

        if not config["dry_run"]:
            if config["db_mode"] == "per_file":
                prepare_database_file(db_path, config["replace_db"])
            elif not prepared_single_db:
                prepare_database_file(db_path, config["replace_db"])
                prepared_single_db = True

        result_db_path, result_table_name, row_count = import_workbook(
            excel_path=excel_file,
            db_path=db_path,
            sheet_name=config["sheet_name"],
            table_name=table_name,
            header_row=int(config["header_row"]),
            data_start_row=int(config["data_start_row"]),
            date_column=config["date_column"],
            dry_run=bool(config["dry_run"]),
            config=config,
        )
        action = "Would create" if config["dry_run"] else "Created"
        print(f"{action}: {result_db_path} | table={result_table_name} | rows={row_count}")
    return 0


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    config = load_config(args.config)
    merged = apply_cli_overrides(config, args)
    return run_import(merged)


if __name__ == "__main__":
    raise SystemExit(main())
